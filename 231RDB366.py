from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from datetime import datetime
from PIL import Image, ImageTk
import io
import tkinter as tk
from urllib.request import urlopen
import re
import requests
from io import BytesIO
import time

def fetch_and_display():
    global panels, driver, wait, mode

    for panel in panels:
        panel.destroy()
    panels.clear()

    wait.until(EC.element_to_be_clickable((By.ID, "stcky"))).click()

    perk_texts = []
    perk_images = []

    for i in range(4):
        text_element = wait.until(EC.visibility_of_element_located((By.ID, f"pn{i}")))
        perk_texts.append(text_element.text)
        image_element = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, f"div.perk_icon#pi{i}")))
        style = image_element.get_attribute('style')
        image_url = re.search(r'url\("(.+?)"\)', style).group(1)
        image_data = urlopen(driver.current_url.rsplit('/', 1)[0] + '/' + image_url).read()
        image = Image.open(io.BytesIO(image_data))
        perk_images.append(image)

    current_time = datetime.now().strftime("%d-%m-%Y | %H:%M")
    max_row = sheet.max_row

    if all(sheet.cell(row=max_row, column=i+1).value is None for i in range(6)):
        start_row = max_row
    else:
        start_row = max_row + 1

    for i, text in enumerate(perk_texts):
        sheet.cell(row=start_row, column=i+1, value=text)
    sheet.cell(row=start_row, column=5, value=mode)
    sheet.cell(row=start_row, column=6, value=current_time)
    workbook.save(excel_file)

    for i, (image, text) in enumerate(zip(perk_images, perk_texts)):
        tk_image = ImageTk.PhotoImage(image)
        panel = tk.Label(root, image=tk_image, text=text, compound="top")
        panel.image = tk_image
        panel.pack(side="left", padx=10, pady=10)
        panels.append(panel)


def init_valorant_agents_mode():
    global driver, wait, mode

    mode = "Valorant Agents"

    valorant_reroll_button.config(command=fetch_valorant_agent)

    hide_main_menu()
    show_reroll_back()

    reroll_button.config(command=fetch_valorant_agent)

    options = Options()
    options.headless = True
    driver = webdriver.Chrome(options=options)
    wait = WebDriverWait(driver, 10)
    driver.get("https://www.valorantpicker.com/#/")

    fetch_valorant_agent()


def fetch_valorant_agent():
    global panels, driver, wait

    for panel in panels:
        panel.destroy()
    panels.clear()

    wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "random-agent-button"))).click()
    
    time.sleep(1)

    wait.until(lambda d: d.find_element(By.CLASS_NAME, "chosen-agent-name").text != "")
    chosen_agent_name = driver.find_element(By.CLASS_NAME, "chosen-agent-name").text

    agent_image_url = driver.find_element(By.CLASS_NAME, "chosen-agent-image").get_attribute('src')
    response = requests.get(agent_image_url)
    img_data = Image.open(BytesIO(response.content))
    tk_image = ImageTk.PhotoImage(img_data)

    agent_name_label = tk.Label(root, text=chosen_agent_name)
    agent_name_label.pack()

    image_label = tk.Label(root, image=tk_image)
    image_label.image = tk_image
    image_label.pack()

    panels.extend([agent_name_label, image_label])

    current_time = datetime.now().strftime("%d-%m-%Y | %H:%M")
    agent_workbook = load_workbook('agentshistory.xlsx')
    agent_sheet = agent_workbook.active
    agent_sheet.append([chosen_agent_name, current_time])
    agent_workbook.save('agentshistory.xlsx')

def init_mode(selected_mode):
    global driver, wait, mode

    mode = selected_mode

    reroll_button.config(command=fetch_and_display)

    hide_main_menu()
    show_reroll_back()

    options = Options()
    options.headless = True
    driver = webdriver.Chrome(options=options)
    wait = WebDriverWait(driver, 10)
    driver.get("https://verewygt.github.io/perkroulette/")

    if mode == "Survivor":
        wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "label[for='surv']"))).click()
    elif mode == "Killer":
        wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "label[for='kill']"))).click()

    fetch_and_display()

def back_to_menu():
    global driver

    for panel in panels:
        panel.destroy()
    panels.clear()

    if 'driver' in globals():
        driver.quit()
        del driver

    hide_reroll_back()
    show_main_menu()

def hide_main_menu():
    survivor_button.pack_forget()
    killer_button.pack_forget()
    valorant_agents_button.pack_forget()

def show_main_menu():
    survivor_button.pack(side="left", padx=10, pady=10)
    killer_button.pack(side="right", padx=10, pady=10)
    valorant_agents_button.pack(side="left", padx=10, pady=10)

def hide_reroll_back():
    reroll_button.pack_forget()
    valorant_reroll_button.pack_forget()
    back_button.pack_forget()

def show_reroll_back():
    if mode == "Valorant Agents":
        valorant_reroll_button.pack(side="bottom", pady=10)
    else:
        reroll_button.pack(side="bottom", pady=10)
    back_button.pack(side="bottom", pady=10)

root = tk.Tk()
root.title("Ingame Randomizer")
panels = []

survivor_button = tk.Button(root, text="Survivor", command=lambda: init_mode("Survivor"))
killer_button = tk.Button(root, text="Killer", command=lambda: init_mode("Killer"))
valorant_agents_button = tk.Button(root, text="Valorant Agents", command=init_valorant_agents_mode)

reroll_button = tk.Button(root, text="Reroll", command=fetch_and_display)
valorant_reroll_button = tk.Button(root, text="Reroll", command=fetch_valorant_agent)
back_button = tk.Button(root, text="Back", command=back_to_menu)

excel_file = 'dbdperks.xlsx'
try:
    workbook = load_workbook(excel_file)
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active

show_main_menu()

root.mainloop()