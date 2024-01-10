import tkinter as tk
from tkinter import ttk
from PIL import Image, ImageTk
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import requests
from io import BytesIO
from openpyxl import load_workbook, Workbook
from datetime import datetime
import time

def fetch_and_update_agent():
    driver.get("https://www.valorantpicker.com/#/")
    random_agent_button = driver.find_element(By.CLASS_NAME, "random-agent-button")
    random_agent_button.click()
    time.sleep(1)
    chosen_agent_name = driver.find_element(By.CLASS_NAME, "chosen-agent-name").text
    agent_image_url = driver.find_element(By.CLASS_NAME, "chosen-agent-image").get_attribute('src')

    response = requests.get(agent_image_url)
    img_data = Image.open(BytesIO(response.content))
    tk_image = ImageTk.PhotoImage(img_data)

    image_label.config(image=tk_image)
    image_label.image = tk_image
    agent_name_label.config(text=chosen_agent_name)

    row = 1
    while sheet.cell(row=row, column=1).value is not None:
        row += 1
    current_time = datetime.now().strftime("%d-%m-%Y | %H:%M")
    sheet.cell(row=row, column=1).value = chosen_agent_name
    sheet.cell(row=row, column=2).value = current_time
    workbook.save(excel_file)

def open_main_window():
    start_window.withdraw()
    global main_window
    main_window = tk.Toplevel()
    setup_main_window()

def setup_main_window():
    global image_label, agent_name_label

    main_window.title("Valorant Agent Picker")

    image_label = ttk.Label(main_window)
    image_label.pack()
    agent_name_label = ttk.Label(main_window, text="")
    agent_name_label.pack()

    reroll_button = ttk.Button(main_window, text="Reroll", command=fetch_and_update_agent)
    reroll_button.pack()

    back_button = ttk.Button(main_window, text="Back", command=back_to_start)
    back_button.pack()

    fetch_and_update_agent()

def back_to_start():
    main_window.destroy()
    start_window.deiconify()

def start_window():
    global start_window
    start_window = tk.Tk()
    start_window.title("Start Window")

    open_button = ttk.Button(start_window, text="Valorant Agents", command=open_main_window)
    open_button.pack()

    start_window.mainloop()

chrome_options = Options()
chrome_options.add_argument("--headless")
driver = webdriver.Chrome(options=chrome_options)

excel_file = 'agentshistory.xlsx'
try:
    workbook = load_workbook(excel_file)
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active

start_window()

driver.quit()
