from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from datetime import datetime
from PIL import Image, ImageTk
import io
import tkinter as tk
from urllib.request import urlopen
import re

def fetch_and_display():
    global panels, driver, wait, mode  # Add 'mode' to the global variables

    # Close panels from previous session if any
    for panel in panels:
        panel.destroy()
    panels.clear()

    # Click the 'Reroll' button
    wait.until(EC.element_to_be_clickable((By.ID, "stcky"))).click()

    perk_texts = []
    perk_images = []
    for i in range(4):
        text_element = wait.until(EC.visibility_of_element_located((By.ID, f"pn{i}")))
        perk_texts.append(text_element.text)

        image_element = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, f"div.perk_icon#pi{i}")))
        style = image_element.get_attribute('style')
        image_url = re.search(r'url\("(.+?)"\)', style).group(1)
        image_data = urlopen(driver.current_url.rsplit('/', 1)[0] + '/' + image_url).read()
        image = Image.open(io.BytesIO(image_data))
        perk_images.append(image)

    current_time = datetime.now().strftime("%d-%m-%Y | %H:%M")
    max_row = sheet.max_row
    if all(sheet.cell(row=max_row, column=i+1).value is None for i in range(6)):
        start_row = max_row
    else:
        start_row = max_row + 1

    for i, text in enumerate(perk_texts):
        sheet.cell(row=start_row, column=i+1, value=text)
    sheet.cell(row=start_row, column=5, value=mode)  # Record 'Survivor' or 'Killer' in the 5th column
    sheet.cell(row=start_row, column=6, value=current_time)  # Record the current time in the 6th column

    workbook.save(excel_file)

    for i, (image, text) in enumerate(zip(perk_images, perk_texts)):
        tk_image = ImageTk.PhotoImage(image)
        panel = tk.Label(root, image=tk_image, text=text, compound="top")
        panel.image = tk_image
        panel.pack(side="left", padx=10, pady=10)
        panels.append(panel)

def init_mode(selected_mode):
    global driver, wait, mode  # Add 'mode' to the global variables

    mode = selected_mode  # Set the mode based on the selected option

    # Hide main menu buttons and show reroll/back buttons
    hide_main_menu()
    show_reroll_back()

    options = Options()
    options.headless = True
    driver = webdriver.Chrome(options=options)
    wait = WebDriverWait(driver, 10)
    driver.get("https://verewygt.github.io/perkroulette/")

    # Click the appropriate label based on the mode
    if mode == "Survivor":
        wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "label[for='surv']"))).click()
    elif mode == "Killer":
        wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "label[for='kill']"))).click()

    fetch_and_display()

def back_to_menu():
    global driver

    # Clear existing images (if any)
    for panel in panels:
        panel.destroy()
    panels.clear()

    # Quit the driver and clean up
    if 'driver' in globals():
        driver.quit()
        del driver

    # Hide reroll/back buttons and show main menu
    hide_reroll_back()
    show_main_menu()

def hide_main_menu():
    survivor_button.pack_forget()
    killer_button.pack_forget()

def show_main_menu():
    survivor_button.pack(side="left", padx=10, pady=10)
    killer_button.pack(side="right", padx=10, pady=10)

def hide_reroll_back():
    reroll_button.pack_forget()
    back_button.pack_forget()

def show_reroll_back():
    reroll_button.pack(side="bottom", pady=10)
    back_button.pack(side="bottom", pady=10)

# Initialization of tkinter GUI
root = tk.Tk()
root.title("DBD Perks")
panels = []

# Main menu buttons
survivor_button = tk.Button(root, text="Survivor", command=lambda: init_mode("Survivor"))
killer_button = tk.Button(root, text="Killer", command=lambda: init_mode("Killer"))

# Reroll and back buttons
reroll_button = tk.Button(root, text="Reroll", command=fetch_and_display)
back_button = tk.Button(root, text="Back", command=back_to_menu)

# Excel file setup
excel_file = 'dbdperks.xlsx'
try:
    workbook = load_workbook(excel_file)
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active

show_main_menu()

root.mainloop()
